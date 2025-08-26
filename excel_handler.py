import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        self._criar_arquivo_se_nao_existir()

    def _criar_arquivo_se_nao_existir(self):
        """Cria o ficheiro Excel com aba 'Pedidos' e cabeçalhos, se não existir."""
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedidos"
            ws.append([
                "Data do Pedido",
                "Nome",
                "Contacto",
                "Produto",
                "Quantidade",
                "Data de Entrega",
                "Observações"
            ])
            wb.save(self.filename)

    def registrar_pedido(self, nome, contacto, produto, quantidade, data_entrega, obs=""):
        """Registra um novo pedido no Excel, criando aba 'Pedidos' se necessário."""
        wb = openpyxl.load_workbook(self.filename)

        # Criar aba "Pedidos" se não existir
        if "Pedidos" not in wb.sheetnames:
            ws = wb.create_sheet("Pedidos")
            ws.append([
                "Data do Pedido",
                "Nome",
                "Contacto",
                "Produto",
                "Quantidade",
                "Data de Entrega",
                "Observações"
            ])
        else:
            ws = wb["Pedidos"]

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Data do registo
            nome,
            contacto,
            produto,
            quantidade,
            data_entrega,
            obs
        ])

        wb.save(self.filename)
